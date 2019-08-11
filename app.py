from openpyxl import Workbook
wb = Workbook()
ws1 = wb.create_sheet("Class")
ws1['A1']= 'nursery'
ws1['A2']= 'jkg'
ws1['A3']= 'skg'
for r in range(4,15):
    for c in range(1,5):
        if c==1:
            section=str(r-3)
        elif c==2:
            section="A"
        elif c==3:
            section="B" 
        else:
            section ="C"       

            
        d= ws1.cell(row=r,column=c,value=section)




# ws2 = wb.create_chartsheet("Name")
